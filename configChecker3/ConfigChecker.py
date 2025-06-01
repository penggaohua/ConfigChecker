#coding=utf-8
import queue
import time
import tkinter as tk
from tkinter import *
from tkinter import ttk,scrolledtext






import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
xlsx_path = r"D:\Config"
#from multiprocessing import Process, Queue

from threading import Thread

#from GUIMain import GUIMain

class ConfigChecker:
    def __init__(self,xlsx_path,xlsx_name,xlsx_sheet):
        print("init")
        print(xlsx_path,xlsx_name,xlsx_sheet)
        self.path = xlsx_path
        self.xlsx_name= xlsx_name
        self.xlsx_sheet= xlsx_sheet
        self.table_title_lines = 4;#表头所在行

        self.xl  = openpyxl.load_workbook(self.path+"\\"+self.xlsx_name)
        self.sheet =self.xl[xlsx_sheet]
        self.none_cell_index=[]
        self.tableTitle_columnLetter_dit={}  #表头和表列字母组成的字典

        for cell in self.sheet[self.table_title_lines]:
            if cell.value is not None:
                self.tableTitle_columnLetter_dit[cell.value] = cell.column_letter


    #得到表头列表
    def get_table_title_list(self):
        return self.get_row_or_column_list(self.table_title_lines)


    #得到某一行或某一列的数据
    def get_row_or_column_list(self,index):
      #参数是数字得到的就是行数据，参数是字母得到的就是列数据
      return [ cell.value for cell in self.sheet[index]]

    #判断某一行是否为空
    def is_row_null(self,row_index):
        res =  self.get_row_or_column_list(row_index)
        return all(x is None for x  in res )

    # 通过index得到列letter
    def get_column_letter_by_index(self,index):
       pass


    #检查当前列非空
    def check_none_column(self,column_name):
        print("开始单列非空检查")
        if(column_name is None):
            return "字段名为空"
        print(self.tableTitle_columnLetter_dit)
        res=""
        column_letter  = self.tableTitle_columnLetter_dit[column_name]

        for i ,cell in enumerate(self.sheet[column_letter]):
            #表头以上的行不检查
            if i+1<self.table_title_lines:
                continue
            print(cell.value)
            if cell.value is None  and self.is_row_null(i+1)== False:
              #  print('{}表\t{}页\t{}列{}列\t第{}行是空\n'.format(self.xlsx_name,self.xlsx_sheet,column_letter,column_name,i+1))
                res += ('{}表\t{}页\t{}列{}\t第{}行是空\n'.format(self.xlsx_name,self.xlsx_sheet,column_letter,column_name,i+1))
        if res=="":
            res = "未发现指定字段列存在空"
        return  res

    #检查当前页sheet非空
    def check_none_sheet(self,sheet_name,q):
        current_sheet =  self.xl[sheet_name]
        table_title_list =  self.get_row_or_column_list(self.table_title_lines) #表头是可能重复或为空
        print(f"表头列表{table_title_list} ")
        res =""
        for index_col,col in enumerate(current_sheet.iter_cols(min_col=1,values_only=True )):
           # if(table_title_list[index_col] is None): continue  #表头为空，这一列就不用检查了todo

            arr = np.array(col)
            none_index_list = np.where(arr == None)[0].tolist()#得到某一列数据，然后把为空的数据对应的index提出来
            #print(none_index_list)
            column_letter = get_column_letter(index_col + 1)
            field_name =  table_title_list[index_col]
            for index  in none_index_list:
                if index + 1 < self.table_title_lines:continue  #表头以上的行不用检查
                title = current_sheet.title
                if(self.is_row_null(index+1) is False):#判断对应的行是否是全空，如果空就跳过了
                    res="{}表\t{}页\t{}列{}\t{}行 为空\n".format(self.xlsx_name,title,column_letter,field_name,index+1)
                    print(res)
                    q.put(res)
                   # print(q)
                    #GUIMain.show_in_text(res, text_ouput)

        #return res



    #检查当前整个表非空
    def check_none_table(self,GUIMain,text_ouput,ignore_list="Sheet1"):
        #得到所有页名
        sheet_names_list = self.xl.sheetnames
       # print(sheet_names_list)
        res =""
        for sheet  in sheet_names_list:
         #   if sheet  in  ignore_list:  continue
           # res += self.check_none_sheet(sheet)
            res = self.check_none_sheet(sheet)
            GUIMain.show_in_text(res,text_ouput)
        #print(res)
        #return res


    def check_test(self):
        print("test")

def multi_test_show(q):
    while True:
        print("打印q的值")
        print(q.get())

def mul_start():
    producer_t = Thread(target=xlrd.check_none_sheet, args=(xlsx_sheet, q,))
    consumer_t = Thread(target=guiMain.show_in_text_process, args=(q, text_output))
    producer_t.start()
    consumer_t.start()
    producer_t.join()
    consumer_t.join()


if __name__ == '__main__':

    guiMain = GUIMain()
    q =queue.Queue()

    # def mul_start():
    #     producer_t = Process(target=xlrd.check_none_sheet, args=(xlsx_sheet, q,))
    #     consumer_t = Process(target=guiMain.show_in_text_process, args=(q, text_output))
    #     producer_t.start()
    #     consumer_t.start()
    #     producer_t.join()
    #     consumer_t.join()


    frame1 = guiMain.set_frame()

    button_test =guiMain.set_button(frame1,"test",mul_start)
    text_output = guiMain.set_scroll_text(frame1, 130, 40)
    button_test.grid(row=0, column=1, sticky=W, pady=10, padx=10)
    text_output.grid(row=0, column=0, sticky=W, pady=10, padx=10)
    frame1.grid(row=4, column=0, pady=5, sticky=N)





    path=  "D:\\Config"
    xlsx_name  = "竞技场神兽配置表.xlsx"
   # xlsx_sheet = "WKshenshou"
    xlsx_sheet = "HCZBShenShou"
   # xlsx_sheet = 'ShenShouBaseAttr'
    field_name = "ShenShouID"

    xlrd = ConfigChecker(path,xlsx_name,xlsx_sheet)
    res = xlrd.get_row_or_column_list(10)
    print(res)


    guiMain.main()





#全表检查
    # res4 =  xlrd.check_none_table()
    # print(res4)



 #   res2 = xlrd.check_none_column(field_name)

  #  res3 = xlrd.check_none_sheet(xlsx_sheet)
    #  xlrd.check_none_table()
  #  print(res3)
   # print(res2)
   # print(xlrd.get_row_or_column_list(248))
#    print(xlrd.is_row_null(248))
   # xlrd.check_none_cell("HP")
